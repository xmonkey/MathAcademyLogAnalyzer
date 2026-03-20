"""Course progress data parser and Excel exporter."""

import re
import json
import pandas as pd
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .pdf_parser import PDFParser


class CourseProgressParser:
    """Parser for course progress data from PDF files."""

    def __init__(self, pdf_path: str):
        """Initialize the course progress parser."""
        self.pdf_parser = PDFParser(pdf_path)

    def parse_course_data(self) -> List[Dict[str, Any]]:
        """Parse course progress data from PDF with improved multiline description handling."""
        text_content = self.pdf_parser.extract_text()

        # Split content by lines for processing
        lines = text_content.split('\n')

        course_data = []
        current_date = None
        daily_total = None
        prefix_lines = []  # Store lines that come before the main data line

        # Improved regex patterns
        date_pattern = r'(\w+),\s+(\w+)\s+(\d{1,2})(?:st|nd|rd|th)?,\s+(\d{4})\s+\((\d+)\s+XP\)'
        # Regular task pattern (e.g., "4th Grade Math Quiz Quiz 1 18/15 XP")
        # The description part is optional - can be "18/18 XP" or "Some Description 18/18 XP"
        regular_task_pattern = r'^(.+?)\s+(Quiz|Lesson|Review|Multistep)\s+(.+)$'
        # Placement task pattern (e.g., "4th Grade Math Placement 35/ XP")
        placement_task_pattern = r'^(.+?)\s+(Placement)\s+(\d+)/\s+XP$'

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            if not line:
                i += 1
                continue

            # Match date line with daily total
            date_match = re.search(date_pattern, line)
            if date_match:
                weekday, month, day, year, total_xp = date_match.groups()
                current_date = f"{year}-{self._month_to_num(month)}-{int(day):02d}"
                daily_total = int(total_xp)
                prefix_lines = []  # Reset prefix lines on new date
                i += 1
                continue

            # Skip header lines and student info
            if (line.startswith('Activity Log') or
                'Student ID:' in line or
                'Start Date:' in line or
                'End Date:' in line or
                line == 'COURSE TASK DESCRIPTION XP EARNED'):
                i += 1
                continue

            # Try to match as a regular task line
            regular_match = re.search(regular_task_pattern, line)
            if regular_match and current_date:
                # Extract course name and description differently
                # Find where the task type is and split accordingly
                task_pattern = r'^(.+?)\s+(Quiz|Lesson|Review|Multistep)\s+(.+)$'
                task_match = re.search(task_pattern, line)
                if task_match:
                    course_part = task_match.group(1).strip()
                    task_type = task_match.group(2)
                    desc_with_xp = task_match.group(3)

                    # Extract XP from the end
                    # Pattern might be "18/18 XP" or "Some Description 18/18 XP"
                    xp_pattern = r'^(.*?)?(\d+)/(\d+)\s+XP$'
                    xp_match = re.search(xp_pattern, desc_with_xp)
                    if xp_match:
                        description = (xp_match.group(1) or '').strip()
                        earned = xp_match.group(2)
                        possible = xp_match.group(3)

                        # Combine with prefix lines if any
                        if prefix_lines:
                            full_description = ' '.join(prefix_lines) + ' ' + description
                            full_description = ' '.join(full_description.split())  # Normalize spaces
                        else:
                            full_description = description

                        # Check if the next line is a continuation (after the XP line)
                        # Look ahead to see if there's a continuation line
                        continuation_lines = []
                        j = i + 1
                        while j < len(lines):
                            next_line = lines[j].strip()
                            # Stop if we hit a new date, empty line, or another XP line
                            if (not next_line or
                                re.search(date_pattern, next_line) or
                                next_line == 'COURSE TASK DESCRIPTION XP EARNED' or
                                re.search(r'\d+/\d+\s+XP', next_line) or
                                re.search(r'\d+/\s+XP', next_line)):
                                break
                            # Check if it looks like text content
                            if re.search(r'[a-zA-Z]{3,}', next_line):
                                continuation_lines.append(next_line)
                                j += 1
                            else:
                                break

                        # Add continuation lines to description
                        if continuation_lines:
                            full_description = full_description + ' ' + ' '.join(continuation_lines)
                            full_description = ' '.join(full_description.split())  # Normalize spaces
                            i = j - 1  # Skip the continuation lines we processed

                        course_data.append({
                            'Date': current_date,
                            'Course': course_part,
                            'Task': task_type.capitalize(),
                            'Description': full_description,
                            'XP Earned': int(earned),
                            'XP Possible': int(possible),
                            'XP Detail': f"{earned}/{possible}",
                            'Daily Total XP': daily_total
                        })
                        prefix_lines = []
                        i += 1
                        continue

            # Try to match as a placement task line
            placement_match = re.search(placement_task_pattern, line)
            if placement_match and current_date:
                course_full, task_type, earned = placement_match.groups()

                # For placement tasks, the XP possible is the same as earned (no slash)
                xp_earned = int(earned)
                xp_possible = xp_earned

                course_data.append({
                    'Date': current_date,
                    'Course': course_full.strip(),
                    'Task': 'Placement',
                    'Description': 'Placement Test',
                    'XP Earned': xp_earned,
                    'XP Possible': xp_possible,
                    'XP Detail': f"{xp_earned}/",
                    'Daily Total XP': daily_total
                })
                prefix_lines = []
                i += 1
                continue

            # If line doesn't match any pattern and isn't a header/date,
            # it might be a continuation line (part of description)
            # Check if it looks like text content (has letters but no XP pattern)
            if line and not re.search(r'\d+/\d+\s+XP', line) and not re.search(r'\d+/\s+XP', line):
                # Check if it looks like a text line (has letters)
                if re.search(r'[a-zA-Z]{3,}', line):
                    prefix_lines.append(line)
                    i += 1
                    continue

            # If we get here, clear prefix lines and move on
            prefix_lines = []
            i += 1

        return course_data

    def export_to_excel(self, output_path: str, data: Optional[List[Dict[str, Any]]] = None):
        """Export course data to Excel with improved formatting."""
        if data is None:
            data = self.parse_course_data()
        
        if not data:
            print("No course data found to export.")
            return
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Sort by date
        df = df.sort_values('Date')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Course Progress"
        
        # Write headers
        headers = [
            'Date', 'Course', 'Task', 'Description', 
            'XP Earned', 'XP Possible', 'XP Detail', 'Daily Total XP'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, (key, value) in enumerate(row_data.items(), 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if key in ['XP Earned', 'XP Possible', 'Daily Total XP']:
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        self._add_summary_sheet(wb, data)
        
        # Save the workbook
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
        print(f"Total records: {len(data)}")
        
        # Print summary
        self._print_summary(data)

    def export_to_json(self, output_path: str, data: Optional[List[Dict[str, Any]]] = None):
        """Export course data to JSON file for data analysis."""
        if data is None:
            data = self.parse_course_data()
        
        if not data:
            print("No course data found to export.")
            return
        
        # Create enhanced JSON structure with metadata
        json_data = {
            "metadata": {
                "total_records": len(data),
                "export_timestamp": datetime.now().isoformat(),
                "source_file": str(self.pdf_parser.pdf_path)
            },
            "statistics": self._calculate_statistics(data),
            "activities": data
        }
        
        # Write to JSON file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        print(f"JSON file saved to: {output_path}")
        print(f"Total records: {len(data)}")
        
        # Print summary
        self._print_summary(data)

    def export_to_excel_and_json(self, excel_path: str, json_path: str, data: Optional[List[Dict[str, Any]]] = None):
        """Export course data to both Excel and JSON formats."""
        if data is None:
            data = self.parse_course_data()
        
        if not data:
            print("No course data found to export.")
            return
        
        # Export to Excel
        self.export_to_excel(excel_path, data)
        
        # Export to JSON
        self.export_to_json(json_path, data)
        
        print(f"Exported to both formats:")
        print(f"  Excel: {excel_path}")
        print(f"  JSON: {json_path}")

    def _calculate_statistics(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate comprehensive statistics for JSON export."""
        if not data:
            return {}
        
        total_xp = sum(item['XP Earned'] for item in data)
        total_possible = sum(item['XP Possible'] for item in data)
        
        # Course breakdown
        courses = {}
        for item in data:
            course = item['Course']
            if course not in courses:
                courses[course] = {'count': 0, 'xp': 0, 'tasks_by_type': {}}
            courses[course]['count'] += 1
            courses[course]['xp'] += item['XP Earned']
            
            # Track tasks by type
            task_type = item['Task']
            if task_type not in courses[course]['tasks_by_type']:
                courses[course]['tasks_by_type'][task_type] = 0
            courses[course]['tasks_by_type'][task_type] += 1
        
        # Task type breakdown
        task_types = {}
        for item in data:
            task_type = item['Task']
            if task_type not in task_types:
                task_types[task_type] = {'count': 0, 'xp': 0}
            task_types[task_type]['count'] += 1
            task_types[task_type]['xp'] += item['XP Earned']
        
        # Daily breakdown
        daily_stats = {}
        for item in data:
            date = item['Date']
            if date not in daily_stats:
                daily_stats[date] = {'total_xp': 0, 'task_count': 0, 'courses': set()}
            daily_stats[date]['total_xp'] += item['XP Earned']
            daily_stats[date]['task_count'] += 1
            daily_stats[date]['courses'].add(item['Course'])
        
        # Convert sets to lists for JSON serialization
        for date_stat in daily_stats.values():
            date_stat['courses'] = list(date_stat['courses'])
        
        return {
            'total_xp': total_xp,
            'total_possible': total_possible,
            'completion_rate': f"{(total_xp/total_possible*100):.1f}%" if total_possible > 0 else '0%',
            'average_xp_per_task': round(total_xp / len(data), 1) if data else 0,
            'unique_courses': len(courses),
            'unique_task_types': len(task_types),
            'date_range': {
                'start': min(item['Date'] for item in data) if data else None,
                'end': max(item['Date'] for item in data) if data else None
            },
            'course_breakdown': courses,
            'task_type_breakdown': task_types,
            'daily_breakdown': daily_stats
        }

    def _add_summary_sheet(self, workbook, data):
        """Add a summary sheet with statistics."""
        summary_ws = workbook.create_sheet("Summary")
        
        # Calculate statistics
        total_xp = sum(item['XP Earned'] for item in data)
        total_possible = sum(item['XP Possible'] for item in data)
        avg_xp_per_task = total_xp / len(data) if data else 0
        
        # Summary data
        summary_data = [
            ['Metric', 'Value'],
            ['Total Records', len(data)],
            ['Total XP Earned', total_xp],
            ['Total XP Possible', total_possible],
            ['Overall Completion Rate', f"{(total_xp/total_possible*100):.1f}%" if total_possible > 0 else '0%'],
            ['Average XP per Task', f"{avg_xp_per_task:.1f}"],
            ['Date Range', f"{min(item['Date'] for item in data)} to {max(item['Date'] for item in data)}" if data else 'N/A'],
            ['Unique Courses', len(set(item['Course'] for item in data))],
        ]
        
        # Write summary
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # Header
                    cell.font = Font(bold=True)
        
        # Course breakdown
        courses = {}
        for item in data:
            course = item['Course']
            if course not in courses:
                courses[course] = {'count': 0, 'xp': 0}
            courses[course]['count'] += 1
            courses[course]['xp'] += item['XP Earned']
        
        # Add course breakdown
        summary_ws.append([])  # Empty row
        summary_ws.append(['Course', 'Tasks', 'Total XP'])
        
        for course, stats in sorted(courses.items()):
            summary_ws.append([course, stats['count'], stats['xp']])

    def _print_summary(self, data):
        """Print a summary of the parsed data."""
        if not data:
            return
        
        total_xp = sum(item['XP Earned'] for item in data)
        courses = set(item['Course'] for item in data)
        task_types = set(item['Task'] for item in data)
        
        print(f"\n=== Summary ===")
        print(f"Total records: {len(data)}")
        print(f"Total XP earned: {total_xp}")
        print(f"Unique courses: {len(courses)}")
        print(f"Task types: {', '.join(task_types)}")
        print(f"Date range: {min(item['Date'] for item in data)} to {max(item['Date'] for item in data)}")

    def _month_to_num(self, month: str) -> str:
        """Convert month name to number."""
        month_map = {
            'January': '01', 'February': '02', 'March': '03', 'April': '04',
            'May': '05', 'June': '06', 'July': '07', 'August': '08',
            'September': '09', 'October': '10', 'November': '11', 'December': '12',
            'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
            'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
            'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
        }
        return month_map.get(month, '01')

    