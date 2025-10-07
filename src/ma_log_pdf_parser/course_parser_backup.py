"""Course progress data parser and Excel exporter."""

import re
import pandas as pd
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .pdf_parser import PDFParser


class CourseProgressParser:
    """Parser for course progress data from PDF files."""

    def __init__(self, pdf_path: str):
        """Initialize the course progress parser."""
        self.pdf_parser = PDFParser(pdf_path)

    def parse_course_data(self) -> List[Dict[str, Any]]:
        """Parse course progress data from PDF with improved logic."""
        text_content = self.pdf_parser.extract_text()
        
        # Split content by lines for processing
        lines = text_content.split('\n')
        
        course_data = []
        current_date = None
        daily_total = None
        
        # Improved regex patterns
        date_pattern = r'(\w+),\s+(\w+)\s+(\d{1,2})(?:st|nd|rd|th)?,\s+(\d{4})\s+\((\d+)\s+XP\)'
        # Regular task pattern (e.g., "4th Grade Math Quiz Quiz 1 18/15 XP")
        regular_task_pattern = r'(.+?)\s+(Quiz|Lesson|Review|Multistep)\s+(.+?)\s+(\d+)/(\d+)\s+XP'
        # Placement task pattern (e.g., "4th Grade Math Placement 35/ XP")
        placement_task_pattern = r'(.+?)\s+(Placement)\s+(\d+)/\s+XP'
        
        for line in lines:
            line = line.strip()
            
            if not line:
                continue
            
            # Match date line with daily total
            date_match = re.search(date_pattern, line)
            if date_match:
                weekday, month, day, year, total_xp = date_match.groups()
                current_date = f"{year}-{self._month_to_num(month)}-{int(day):02d}"
                daily_total = int(total_xp)
                continue
            
            # Skip header lines and student info
            if (line.startswith('Activity Log') or 
                'Student ID:' in line or 
                'Start Date:' in line or 
                'End Date:' in line or
                line == 'COURSE TASK DESCRIPTION XP EARNED'):
                continue
            
            # Match regular course task lines
            regular_match = re.search(regular_task_pattern, line)
            if regular_match and current_date:
                course_full, task_type, description, earned, possible = regular_match.groups()
                
                # Extract course name
                course = self._extract_course_name(course_full)
                
                # Use task_type directly for Task
                task_name = task_type.capitalize()
                
                course_data.append({
                    'Date': current_date,
                    'Course': course,
                    'Task': task_name,
                    'Description': description.strip(),
                    'XP Earned': int(earned),
                    'XP Possible': int(possible),
                    'XP Detail': f"{earned}/{possible}",
                    'Daily Total XP': daily_total
                })
                continue
            
            # Match placement task lines
            placement_match = re.search(placement_task_pattern, line)
            if placement_match and current_date:
                course_full, task_type, earned = placement_match.groups()
                
                # Extract course name
                course = self._extract_course_name(course_full)
                
                # For placement tasks, the XP possible is the same as earned (no slash)
                xp_earned = int(earned)
                xp_possible = xp_earned
                
                course_data.append({
                    'Date': current_date,
                    'Course': course,
                    'Task': 'Placement',
                    'Description': 'Placement Test',
                    'XP Earned': xp_earned,
                    'XP Possible': xp_possible,
                    'XP Detail': f"{xp_earned}/",
                    'Daily Total XP': daily_total
                })
        
        return course_data

    def export_to_excel(self, output_path: str, data: Optional[List[Dict[str, Any]]] = None):
        """Export course data to Excel with improved formatting."""
        if data is None:
            data = self.parse_course_data()
        
        if not data:
            print("No course data found to export.")
            return
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Sort by date
        df = df.sort_values('Date')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Course Progress"
        
        # Write headers
        headers = [
            'Date', 'Course', 'Task', 'Description', 
            'XP Earned', 'XP Possible', 'XP Detail', 'Daily Total XP'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, (key, value) in enumerate(row_data.items(), 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if key in ['XP Earned', 'XP Possible', 'Daily Total XP']:
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        self._add_summary_sheet(wb, data)
        
        # Save the workbook
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
        print(f"Total records: {len(data)}")
        
        # Print summary
        self._print_summary(data)

    def _add_summary_sheet(self, workbook, data):
        """Add a summary sheet with statistics."""
        summary_ws = workbook.create_sheet("Summary")
        
        # Calculate statistics
        total_xp = sum(item['XP Earned'] for item in data)
        total_possible = sum(item['XP Possible'] for item in data)
        avg_xp_per_task = total_xp / len(data) if data else 0
        
        # Summary data
        summary_data = [
            ['Metric', 'Value'],
            ['Total Records', len(data)],
            ['Total XP Earned', total_xp],
            ['Total XP Possible', total_possible],
            ['Overall Completion Rate', f"{(total_xp/total_possible*100):.1f}%" if total_possible > 0 else '0%'],
            ['Average XP per Task', f"{avg_xp_per_task:.1f}"],
            ['Date Range', f"{min(item['Date'] for item in data)} to {max(item['Date'] for item in data)}" if data else 'N/A'],
            ['Unique Courses', len(set(item['Course'] for item in data))],
        ]
        
        # Write summary
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # Header
                    cell.font = Font(bold=True)
        
        # Course breakdown
        courses = {}
        for item in data:
            course = item['Course']
            if course not in courses:
                courses[course] = {'count': 0, 'xp': 0}
            courses[course]['count'] += 1
            courses[course]['xp'] += item['XP Earned']
        
        # Add course breakdown
        summary_ws.append([])  # Empty row
        summary_ws.append(['Course', 'Tasks', 'Total XP'])
        
        for course, stats in sorted(courses.items()):
            summary_ws.append([course, stats['count'], stats['xp']])

    def _print_summary(self, data):
        """Print a summary of the parsed data."""
        if not data:
            return
        
        total_xp = sum(item['XP Earned'] for item in data)
        courses = set(item['Course'] for item in data)
        task_types = set(item['Task'] for item in data)
        
        print(f"\n=== Summary ===")
        print(f"Total records: {len(data)}")
        print(f"Total XP earned: {total_xp}")
        print(f"Unique courses: {len(courses)}")
        print(f"Task types: {', '.join(task_types)}")
        print(f"Date range: {min(item['Date'] for item in data)} to {max(item['Date'] for item in data)}")

    def _month_to_num(self, month: str) -> str:
        """Convert month name to number."""
        month_map = {
            'January': '01', 'February': '02', 'March': '03', 'April': '04',
            'May': '05', 'June': '06', 'July': '07', 'August': '08',
            'September': '09', 'October': '10', 'November': '11', 'December': '12',
            'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
            'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
            'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
        }
        return month_map.get(month, '01')

    def _extract_course_name(self, full_name: str) -> str:
        """Extract clean course name from full name."""
        # Keep the full name but clean it up
        course_name = full_name.strip()
        
        # Don't remove grade level prefixes - keep "4th Grade Math" format
        
        # Capitalize properly
        course_name = course_name.title()
        
        return course_name